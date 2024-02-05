import pandas as pd
import openpyxl
import requests
from openpyxl.utils import column_index_from_string


def login_to_website(url, username, password):
    data = {
        'AUTH_FORM': 'Y',
        'TYPE': 'AUTH',
        'backurl': "/auth/?backurl=%2F",
        'USER_LOGIN': username,
        'USER_PASSWORD': password,
        'USER_REMEMBER': 'Y',
        'Login': 'Войти'
    }

    response = session.post(url, data=data)
    # print(response.status_code)
    # print(response.text)

    # Дополнительная обработка ответа, если необходимо
    if response.status_code == 200:
        print("Успешно вошли на веб-сайт")
    else:
        print("Не удалось войти на веб-сайт")


def download_file(url, file_name):
    response = session.get(url)
    with open(file_name, 'wb') as file:
        file.write(response.content)
    print(f"Файл '{file_name}' успешно загружен")


def remove_rows(file_name, start_row, end_row):
    wb = openpyxl.load_workbook(file_name)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(start_row, end_row)

    wb.save(file_name)
    print(f"Строки {start_row}-{end_row} удалены из всех листов файла '{file_name}'")


def remove_rows_for_one_sheet(file_name, sheet_name):
    # Загрузка файла Excel
    df = pd.read_excel(file_name, sheet_name=sheet_name)

    # Поиск индекса строки, где первая колонка называется "Категория"
    category_row_index = df[df.iloc[:, 0] == 'Категория'].index[0]

    # Удаление строк до строки с категорией
    df = df.iloc[category_row_index:]
    print(f"Строки до -Категория- удалены с листа '{sheet_name}'")

    # Сохранение изменений в файле Excel
    df.to_excel(file_name, sheet_name=sheet_name, index=False, header=False)

    rename_next_column(file_name, sheet_name, "Цена (партн)", "Валюта (партн)")
    rename_next_column(file_name, sheet_name, "Цена (розн)", "Валюта (розн)")


def select_columns(file_name, sheet_name, columns):
    df = pd.read_excel(file_name, sheet_name=sheet_name)
    selected_df = df[columns]
    return selected_df


def rename_next_column(file_name, sheet_name, column_name, new_column_name):
    # Загрузка файла Excel
    df = pd.read_excel(file_name, sheet_name=sheet_name)

    # Получение индекса колонки, после которой нужно переименовать следующую колонку
    column_index = df.columns.get_loc(column_name)

    # Получение имени следующей колонки
    next_column_name = df.columns[column_index + 1]

    # Переименование следующей колонки
    df.rename(columns={next_column_name: new_column_name}, inplace=True)

    # Сохранение изменений в файле Excel
    df.to_excel(file_name, sheet_name=sheet_name, index=False)


def update_prices_with_exchange_rate(file_name, sheet_name, currency_column, price_column):
    # Загрузка файла Excel
    df = pd.read_excel(file_name, sheet_name=sheet_name)

    # Получение текущего курса доллара
    response = requests.get('https://api.exchangerate-api.com/v4/latest/USD')
    data = response.json()
    exchange_rate = data['rates']['RUB']
    print(f"Курс 1 доллара = {exchange_rate} руб.")

    # Обновление цен с помощью курса обмена
    for index, row in df.iterrows():
        if row[currency_column] == 'USD':
            price = row[price_column]
            new_price = round(price * exchange_rate, 2)  # Округляем до 2 знаков после запятой
            df.at[index, price_column] = new_price
            df.at[index, currency_column] = "руб."

    # Сохранение изменений в файле Excel
    df.to_excel(file_name, sheet_name=sheet_name, index=False)


def add_column_to_xlsx(filename):
    # Открываем файл
    wb = openpyxl.load_workbook(filename)
    # Выбираем активный лист
    sheet = wb.active

    # Находим индекс колонки "Доступно к заказу"
    dostupno_k_zakazu_index = None
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == "Доступно к заказу":
                dostupno_k_zakazu_index = cell.column_letter

    # Вставляем новую колонку "FBY" после колонки "Доступно к заказу"
    sheet.insert_cols(column_index_from_string(dostupno_k_zakazu_index) + 1)

    # Назначаем заголовок для новой колонки
    new_column_header = "FBY"
    sheet.cell(row=1, column=column_index_from_string(dostupno_k_zakazu_index) + 1, value=new_column_header)

    # Сохраняем изменения в файле
    wb.save(filename)

    # Открываем файл
    wb = openpyxl.load_workbook(filename)
    # Выбираем активный лист
    sheet = wb.active

    # Получаем максимальное количество строк в файле
    max_row = sheet.max_row

    # Находим индекс колонки "Доступно к заказу"
    dostupno_k_zakazu_index = None
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == "Доступно к заказу":
                dostupno_k_zakazu_index = cell.column
            elif cell.value == "FBY":
                fby = cell.value
                # print(fby)

    # Перебираем данные в колонке "Доступно к заказу"
    for row in range(2, max_row + 1):
        dostupno_k_zakazu = int(sheet.cell(row=row, column=dostupno_k_zakazu_index).value)

        # Проверяем тип данных
        data_type = type(dostupno_k_zakazu)

        # Проверяем значение и записываем в колонку "FBY"
        if dostupno_k_zakazu == 0:
            sheet.cell(row=row, column=dostupno_k_zakazu_index + 1).value = "да"
        else:
            sheet.cell(row=row, column=dostupno_k_zakazu_index + 1).value = None

    # Сохраняем изменения в файле
    wb.save(filename)


def write_to_xlsx_file(file_name, sheet_name, data):
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
    data.to_excel(writer, sheet_name=sheet_name, index=False)
    writer._save()
    print(f"Данные успешно записаны в файл '{file_name}'")


def convert_str_to_float(file_name, sheet_name, column_name):
    # Загрузка файла Excel
    df = pd.read_excel(file_name, sheet_name=sheet_name)

    # Удаление пробелов из значений и преобразование строковых значений в числа
    df[column_name] = df[column_name].apply(lambda x: float(x.replace(',', '.').replace(' ', '')))

    # Сохранение изменений в файле Excel
    df.to_excel(file_name, sheet_name=sheet_name, index=False)


# Пример использования
if __name__ == '__main__':
    session = requests.Session()
    session.verify = True

    url = "https://b2b.digis.ru/"
    username = "info@isintellekt.ru"
    password = "reder378"

    login_to_website(url, username, password)

    href_file = '<a class="ttl" href="/bitrix/redirect.php?event1=news_out&event2=/personal/profile/price/slvhf/daily_price_im_slvhf.xlsx&event3=/personal/profile/price/slvhf/daily_price_im_slvhf.xlsx&goto=/personal/profile/price/slvhf/daily_price_im_slvhf.xlsx">'
    url_file = "https://b2b.digis.ru/bitrix/redirect.php?event1=news_out&event2=/personal/profile/price/slvhf/daily_price_im_slvhf.xlsx&event3=/personal/profile/price/slvhf/daily_price_im_slvhf.xlsx&goto=/personal/profile/price/slvhf/daily_price_im_slvhf.xlsx"

    file_name = "daily_price_im_slvhf.xlsx"
    download_file(url_file, file_name)

    sheet_name = "Объединенный прайс-лист"
    remove_rows_for_one_sheet(file_name, sheet_name)

    currency_column = 'Валюта (партн)'
    price_column = 'Цена (партн)'
    update_prices_with_exchange_rate(file_name, sheet_name, currency_column, price_column)

    currency_column = 'Валюта (розн)'
    price_column = 'Цена (розн)'
    update_prices_with_exchange_rate(file_name, sheet_name, currency_column, price_column)

    columns = ['Категория', 'Подкатегория', 'Бренд', 'Код', 'Артикул', 'Наименование', 'На складе', 'Доступно к заказу', 'Цена (партн)', 'Валюта (партн)', 'Цена (розн)', 'Валюта (розн)']
    selected_data = select_columns(file_name, sheet_name, columns)
    write_to_xlsx_file("output_daily.xlsx", sheet_name, selected_data)
    convert_str_to_float("output_daily.xlsx", "Объединенный прайс-лист", "Доступно к заказу")
    add_column_to_xlsx("output_daily.xlsx")

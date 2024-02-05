import gdown
import openpyxl
import pandas as pd


def download_file(url, output_filename):
    gdown.download(url, output_filename, quiet=False)
    file_name = output_filename.split("/")[-1]

    return file_name


def remove_sheet(file_path, sheet_name):
    # Открываем файл
    wb = openpyxl.load_workbook(file_path)

    # Удаляем лист, если он существует
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        wb.remove(sheet)

    # Сохраняем изменения
    wb.save(file_path)


def select_columns(file_name, sheet_name, price_columns):
    df = pd.read_excel(file_name, sheet_name=sheet_name)

    # Выбираем столбец с ценой в зависимости от того, какой из вариантов присутствует в данных
    selected_price_column = None
    for price_col in price_columns:
        if price_col in df.columns:
            selected_price_column = price_col
            break

    if selected_price_column is None:
        raise ValueError(f"No price column found in the specified options {price_columns} for sheet {sheet_name}.")

    # Выбираем все столбцы, включая выбранный столбец с ценой
    selected_columns = list(df.columns) + [selected_price_column]
    selected_data = df[selected_columns]

    return selected_data


def delete_empty_first_row(file_path):
    # Открываем файл
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Проверяем, что лист не пустой и первая строка не пустая
        if sheet.max_row > 0 and all(cell.value is None for cell in sheet[1]):
            # Удаляем первую строку
            sheet.delete_rows(1, 1)

    # Сохраняем изменения
    wb.save(file_path)


def write_to_xlsx_file(file_name, sheet_name, data):
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
    data.to_excel(writer, sheet_name=sheet_name, index=False)
    writer._save()
    print(f"Данные успешно записаны в файл '{file_name}'")


if __name__ == '__main__':
    file_name = download_file("https://docs.google.com/spreadsheets/d/1zlz3u_k_RFuPVkPjQsSUz0qgnQMdzoNChdo9_6Ku9DE/export?format=xlsx", "output_bureau.xlsx")

    try:
       remove_sheet("output_bureau.xlsx", "Новости")
    except:
        print("Лист Новости не найден")

    delete_empty_first_row(file_name)

    wb = openpyxl.load_workbook(file_name)
    sheet_names = wb.sheetnames
    price_columns = ['РРЦ, руб. (актуальная) ', 'РРЦ, руб.', 'РРЦ, руб. ', 'ДИЛЕР, руб. ', 'РРЦ, руб. ', 'ДИЛЕР, руб.', 'РРЦ, руб. \n']

    for sheet_name in sheet_names:
        print(f"Название листа: {sheet_name}")

        try:
            selected_data = select_columns(file_name, sheet_name, price_columns)
            write_to_xlsx_file("output_bureau_result.xlsx", "Результат", selected_data)

            print(selected_data)
        except ValueError as e:
            print(e)




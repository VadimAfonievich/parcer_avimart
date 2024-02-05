import csv
import openpyxl
import pandas as pd


def open_csv_file(file_name):
    with open(file_name, encoding='cp1251') as csv_file:
        csv_reader = csv.reader(csv_file)
        rows = list(csv_reader)  # Преобразуем итератор в список
        row_count = len(rows)  # Подсчитываем количество строк

        print("Количество столбцов:", len(rows[0]))

        i = 0
        for row in rows:
            # print(row)
            i += 1
        print("Количество строк:", i)


def write_to_xlsx_file(file_name, data):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Запись данных в ячейки
    for row in data:
        ws.append(row)

    # Сохранение файла
    wb.save(file_name)
    print(f"Данные успешно записаны в файл '{file_name}'")


def open_xlsx_file(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    results = []

    # Пример: вывод содержимого всех ячеек
    for row in ws.iter_rows(values_only=True):
        data = list(row[0:6]) + list(row[18:21]) + [row[24]]
        results.append(data)

    del results[:2]

    for result in results:
        print(result)

    # записываем данные первого листа в файл
    write_to_xlsx_file("pars_result_daily.xlsx", results)

    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        print(f"Название листа: {sheet_name}")

        ws = wb[sheet_name]
        # Пример: вывод содержимого всех ячеек
        for row in ws.iter_rows(values_only=True):
            del results[:2]
            data = list(row[0:6]) + list(row[18:21]) + [row[24]]
            results.append(data)

        for result in results:
            print(result)

    # записываем данные первого листа в файл
    write_to_xlsx_file("pars_result_daily.xlsx", results)




def remove_sheet_by_name(file_name, sheet_name):
    wb = openpyxl.load_workbook(file_name)
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        wb.remove(sheet)
        wb.save(file_name)
        print(f"Лист '{sheet_name}' удален из файла '{file_name}'")
    else:
        print(f"Лист '{sheet_name}' не найден в файле '{file_name}'")


def delete_zero_rows(file_name):
    with open(file_name, 'r', encoding='cp1251') as csv_file:
        csv_reader = csv.reader(csv_file)
        rows = [row for row in csv_reader if any(row)]  # Фильтруем пустые строки

        with open(file_name, 'w', newline='', encoding='cp1251') as new_csv_file:
            csv_writer = csv.writer(new_csv_file)
            csv_writer.writerows(rows)  # Записываем отфильтрованные строки в новый файл


# Пример использования
if __name__ == '__main__':
    print("START WORKING")
    # open_csv_file("./01_avimart price list/export.csv")

    # try:
    #     remove_sheet_by_name("./01_avimart price list/daily_price_im_slvhf.xlsx", "Оглавление")
    # except Exception:
    #     print("Листа с названием 'Оглавление' не найдена в прайсе")

    # open_xlsx_file("./01_avimart price list/daily_price_im_slvhf.xlsx")
